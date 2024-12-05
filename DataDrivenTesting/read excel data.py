import openpyxl

# Path to the Excel file
file = "C:\\Users\\Admin\\Desktop\\Tutorial\\Selenium\\DataDrivenTest.xlsx"

# Load the Excel workbook
workBook = openpyxl.load_workbook(file)

# Load a specific sheet by name
sheet = workBook["Sheet1"]

# Get the total number of rows in the sheet
rows = sheet.max_row
print(f"Total rows: {rows}")

# Get the total number of columns in the sheet
columns = sheet.max_column
print(f"Total columns: {columns}")

# Iterate through rows and columns to extract and print cell data
print("Data from the Excel sheet:")
for r in range(1, rows + 1):  # Iterate through each row
    for c in range(1, columns + 1):  # Iterate through each column
        # Print the value of each cell, formatted with tab space
        print(sheet.cell(r, c).value, end='    ')
    print()  # Move to the next line after each row
