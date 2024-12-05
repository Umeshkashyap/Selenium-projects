import openpyxl
file=r"C:\Users\Admin\Desktop\Tutorial\Selenium\\fd.xlsx"

# Load the Excel workbook
workBook = openpyxl.load_workbook(file)

# Load a specific sheet by name
sheet = workBook["Sheet1"]

sheet.cell(2,8).value="heading1"
sheet.cell(3,8).value="heading2"
sheet.cell(4,8).value="heading3"
sheet.cell(5,8).value="heading4"


workBook.save(file)
print(f"Data written to {file} successfully.")