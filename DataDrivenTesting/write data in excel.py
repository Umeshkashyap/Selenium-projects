import openpyxl

# Path to the Excel file
file = "C:\\Users\\Admin\\Desktop\\Tutorial\\Selenium\\writeData.xlsx"

# Load the Excel workbook
workBook = openpyxl.load_workbook(file)

# Load a specific sheet by name
sheet = workBook["Sheet1"]

# Update cells in the specified range for writing the data
# for inserting same type of data in excell
for r in range(1, 6):  # Iterate through rows 1 to 5
    for c in range(1, 4):  # Iterate through columns 1 to 3
        # Set the value of each cell to "Testing"
        sheet.cell(r, c).value = "Testing"

# Save the updated Excel file
workBook.save(file)

# Confirm save
print(f"Data written to {file} successfully.")


#----------------------------------------------------------------------------------
# Path to the Excel file
xlfile = "C:\\Users\\Admin\\Desktop\\Tutorial\\Selenium\\writespecificdata.xlsx"

# Load the Excel workbook
workBook = openpyxl.load_workbook(xlfile)

# Load a specific sheet by name
sheet = workBook["Sheet1"]

sheet.cell(1,1).value="heading1"
sheet.cell(1,2).value="heading2"
sheet.cell(1,3).value="heading3"
sheet.cell(1,4).value="heading4"
sheet.cell(2,1).value="test4"
sheet.cell(2,2).value="test5"
sheet.cell(2,3).value="test6"
sheet.cell(2,4).value="test7"
sheet.cell(3,1).value="tess"
sheet.cell(3,2).value="tt8s"
sheet.cell(3,3).value="test"
sheet.cell(3,3).value="test0"

workBook.save(xlfile)
print(f"Data written to {xlfile} successfully.")
