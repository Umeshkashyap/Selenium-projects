import openpyxl  # Library for working with Excel files
from openpyxl.styles import PatternFill  # For applying color fills to cells

def get_row(file, sheetName):
    # Returns the total number of rows in the specified Excel sheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

def get_column(file, sheetName):
    # Returns the total number of columns in the specified Excel sheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

def read_data(file, sheetName, rowno, columnno):
    # Reads and returns the value from the specified cell
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, columnno).value

def write_Data(file, sheetName, rowno, columnno, data):
    # Writes data to the specified cell in the Excel sheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, columnno).value = data
    workbook.save(file)

def fill_green(file, sheetName, rowno, columnno):
    # Fills the specified cell with green color
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    fillgreen = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rowno, columnno).fill = fillgreen
    workbook.save(file)

def fill_red(file, sheetName, rowno, columnno):
    # Fills the specified cell with red color
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    fillred = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rowno, columnno).fill = fillred
    workbook.save(file)
